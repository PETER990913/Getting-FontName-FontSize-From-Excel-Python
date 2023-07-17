import openpyxl
import pandas as pd
# Load the Excel file
# workbook = openpyxl.load_workbook('data.xlsx')
workbook = openpyxl.load_workbook('new.xlsx')

# Select the desired worksheet
worksheet = workbook['Sheet1']  # Replace 'Sheet1' with the name of your worksheet
# worksheet = workbook['Data']  # Replace 'Sheet1' with the name of your worksheet

# Iterate over each cell in the worksheet
Font_Name_List = []
Font_Size_List = []
for row in worksheet.iter_rows():
    for cell in row:
        # Get the font information for the cell
        font = cell.font
        font_family = font.name
        Font_Name_List.append(font_family)
        font_size = font.size
        Font_Size_List.append(font_size)
        
NoRepeated_Font_Name_List = list(set(Font_Name_List))
NoRepeated_Font_Size_List = list(set(Font_Size_List))
# print(Font_Name_List,len(Font_Name_List))
# print(Font_Size_List, len(Font_Size_List))
# print(NoRepeated_Font_Name_List, len(NoRepeated_Font_Name_List))
# print(NoRepeated_Font_Size_List, len(NoRepeated_Font_Size_List))

Font_Name_Excel_List=[]
Font_Name_Number_Excel_List = []
new_data_name = {}
for k in range(0, len(NoRepeated_Font_Name_List)):
    Font_Name_Number = 0
    for i in range(0, len(Font_Name_List)):        
        if Font_Name_List[i] == f"{NoRepeated_Font_Name_List[k]}":
            Font_Name_Number += 1
    new_data_name[f"{NoRepeated_Font_Name_List[k]}"] = Font_Name_Number
    Font_Name_Excel_List.append(f"{NoRepeated_Font_Name_List[k]}")
    Font_Name_Number_Excel_List.append(Font_Name_Number)
print(new_data_name)
print(Font_Name_Excel_List)
print(Font_Name_Number_Excel_List)
dict = {'Font_Name' : Font_Name_Excel_List, 'Count' : Font_Name_Number_Excel_List}
df = pd.DataFrame(dict)
df.to_csv('result_Font_Name.csv') 

Font_Size_Excel_List=[]
Font_Size_Number_Excel_List = []
new_data_size = {}
for l in range(0, len(NoRepeated_Font_Size_List)):
    Font_Size_Number = 0
    for i in range(0, len(Font_Name_List)):        
        if Font_Size_List[i] == NoRepeated_Font_Size_List[l]:
            Font_Size_Number += 1
    new_data_size[f"{NoRepeated_Font_Size_List[l]}"] = Font_Size_Number
    Font_Size_Excel_List.append(f"{NoRepeated_Font_Size_List[l]}")
    Font_Size_Number_Excel_List.append(Font_Size_Number)
print(new_data_size)
print(Font_Size_Excel_List)
print(Font_Size_Number_Excel_List)
dict = {'Font_Size' : Font_Size_Excel_List, 'Count' : Font_Size_Number_Excel_List}
df = pd.DataFrame(dict)
df.to_csv('result_Font_Size.csv') 


